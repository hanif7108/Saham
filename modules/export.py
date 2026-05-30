"""
Export Module
=============
Generate laporan dari hasil scan/portfolio:
  - Excel (.xlsx) — multi-sheet (Ranking, Portfolio, Alerts, Summary)
  - PDF (executive summary + tabel ringkas)

Library:
  - openpyxl untuk xlsx (sudah dependency umum)
  - reportlab untuk PDF
"""

import os
from datetime import datetime
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SAHAM_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXPORT_DIR = os.path.join(SAHAM_DIR, "exports")


def _ensure_export_dir():
    os.makedirs(EXPORT_DIR, exist_ok=True)


# ----------------------------- EXCEL ----------------------------- #

# Style tokens
HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
GREEN_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
RED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")


def _autosize(ws):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 35)


def _write_header(ws, headers, row=1):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")


def export_xlsx(
    ranking: List[Dict],
    portfolio: List[Dict],
    alerts: List[Dict],
    summary: Dict,
    out_path: str = None
) -> str:
    """Generate Excel workbook multi-sheet."""
    _ensure_export_dir()
    if out_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        out_path = os.path.join(EXPORT_DIR, f"sharia_report_{ts}.xlsx")

    wb = Workbook()

    # ---- Sheet 1: Summary ----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Sharia Trading Assistant — Laporan"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(italic=True, color="666666")

    summary_rows = [
        ["Total emiten ter-scan", summary.get("total_stocks", 0)],
        ["Posisi portfolio", summary.get("portfolio_count", 0)],
        ["Avg P/L portfolio", f"{summary.get('avg_pnl', 0):.2f}%"],
        ["Total alert hari ini", summary.get("alert_count", 0)],
        ["Strong Buy candidates", summary.get("strong_buy_count", 0)],
        ["Oversold gems", summary.get("oversold_count", 0)],
    ]
    for i, (label, value) in enumerate(summary_rows, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    _autosize(ws)

    # ---- Sheet 2: Ranking ----
    ws = wb.create_sheet("Ranking")
    headers = ["#", "Ticker", "Sektor", "Harga", "Fund Score",
               "Stars", "Label", "Tech Signal", "RSI", "MA Cross", "Final Signal"]
    _write_header(ws, headers)
    for i, s in enumerate(ranking, start=1):
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=s.get("ticker"))
        ws.cell(row=row, column=3, value=s.get("sector"))
        ws.cell(row=row, column=4, value=s.get("current_price"))
        ws.cell(row=row, column=5,
                value=f'{s.get("fundamental_score")}/{s.get("fundamental_max")}')
        ws.cell(row=row, column=6, value=s.get("stars"))
        ws.cell(row=row, column=7, value=s.get("fundamental_label"))
        ws.cell(row=row, column=8, value=s.get("technical_signal"))
        ws.cell(row=row, column=9, value=s.get("rsi"))
        ws.cell(row=row, column=10, value=s.get("ma_cross"))

        final = s.get("final_signal")
        c = ws.cell(row=row, column=11, value=final)
        if final in ("STRONG BUY", "BUY"):
            c.fill = GREEN_FILL
        elif final == "AVOID":
            c.fill = RED_FILL
        elif "ACCUMULATE" in (final or "") or "WATCH" in (final or ""):
            c.fill = YELLOW_FILL

    ws.freeze_panes = "A2"
    _autosize(ws)

    # ---- Sheet 3: Portfolio ----
    ws = wb.create_sheet("Portfolio")
    headers = ["Ticker", "Sektor", "Harga Beli", "Harga Sekarang", "Lot",
               "P/L %", "Fund Score", "Tech Signal", "Aksi"]
    _write_header(ws, headers)
    for i, p in enumerate(portfolio, start=1):
        row = i + 1
        ws.cell(row=row, column=1, value=p.get("ticker"))
        ws.cell(row=row, column=2, value=p.get("sector"))
        ws.cell(row=row, column=3, value=p.get("harga_beli"))
        ws.cell(row=row, column=4, value=p.get("sekarang"))
        ws.cell(row=row, column=5, value=p.get("lot"))
        c_pnl = ws.cell(row=row, column=6, value=p.get("pnl"))
        c_pnl.fill = GREEN_FILL if (p.get("pnl") or 0) >= 0 else RED_FILL
        ws.cell(row=row, column=7,
                value=f'{p.get("fundamental_score")}/{p.get("fundamental_max")}')
        ws.cell(row=row, column=8, value=p.get("technical_signal"))
        c_aksi = ws.cell(row=row, column=9, value=p.get("aksi"))
        if "SELL" in (p.get("aksi") or ""):
            c_aksi.fill = RED_FILL
        elif "BUY" in (p.get("aksi") or ""):
            c_aksi.fill = GREEN_FILL

    ws.freeze_panes = "A2"
    _autosize(ws)

    # ---- Sheet 4: Alerts ----
    ws = wb.create_sheet("Alerts")
    headers = ["Tipe", "Severity", "Ticker", "Pesan"]
    _write_header(ws, headers)
    for i, a in enumerate(alerts, start=1):
        row = i + 1
        ws.cell(row=row, column=1, value=a.get("type"))
        c_sev = ws.cell(row=row, column=2, value=a.get("severity"))
        if a.get("severity") == "high":
            c_sev.fill = RED_FILL
        else:
            c_sev.fill = YELLOW_FILL
        ws.cell(row=row, column=3, value=a.get("ticker"))
        ws.cell(row=row, column=4, value=a.get("message"))

    ws.freeze_panes = "A2"
    _autosize(ws)

    wb.save(out_path)
    return out_path


# ----------------------------- PDF ----------------------------- #

def export_pdf(
    ranking: List[Dict],
    portfolio: List[Dict],
    alerts: List[Dict],
    summary: Dict,
    out_path: str = None
) -> str:
    """Generate PDF executive summary."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                     Table, TableStyle)

    _ensure_export_dir()
    if out_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        out_path = os.path.join(EXPORT_DIR, f"sharia_report_{ts}.pdf")

    doc = SimpleDocTemplate(out_path, pagesize=A4,
                             leftMargin=1.5*cm, rightMargin=1.5*cm,
                             topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = []

    # Custom styles for tables to enforce proper text wrapping and sizing
    th_style = ParagraphStyle(
        "TableHeader",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=11,
        textColor=colors.white,
        alignment=1  # Centered
    )

    tc_style = ParagraphStyle(
        "TableCell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.black,
        alignment=0  # Left-aligned
    )

    tcc_style = ParagraphStyle(
        "TableCellCenter",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.black,
        alignment=1  # Centered
    )

    story.append(Paragraph(
        "<b>Sharia Trading Assistant — Laporan Harian</b>",
        ParagraphStyle("title", parent=styles["Title"], fontSize=18)
    ))
    story.append(Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        styles["Italic"]
    ))
    story.append(Spacer(1, 0.5*cm))

    # ---- Summary box ----
    story.append(Paragraph("<b>Ringkasan</b>", styles["Heading2"]))
    summary_data = [
        ["Total emiten", summary.get("total_stocks", 0)],
        ["Portfolio", summary.get("portfolio_count", 0)],
        ["Avg P/L", f"{summary.get('avg_pnl', 0):.2f}%"],
        ["Strong Buy candidates", summary.get("strong_buy_count", 0)],
        ["Oversold gems", summary.get("oversold_count", 0)],
        ["Total alerts", summary.get("alert_count", 0)],
    ]
    t = Table(summary_data, colWidths=[6*cm, 4*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E0E7FF")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # ---- Top 10 Ranking ----
    story.append(Paragraph("<b>Top 10 Emiten Syariah (by Fundamental)</b>",
                            styles["Heading2"]))
    
    rank_data = [[
        Paragraph("<b>#</b>", th_style),
        Paragraph("<b>Ticker</b>", th_style),
        Paragraph("<b>Sektor</b>", th_style),
        Paragraph("<b>Score</b>", th_style),
        Paragraph("<b>RSI</b>", th_style),
        Paragraph("<b>Signal</b>", th_style)
    ]]
    for i, s in enumerate(ranking[:10], 1):
        final_sig = s.get("final_signal", "")
        sig_color = "black"
        if final_sig in ("STRONG BUY", "BUY"):
            sig_color = "#15803D"
        elif final_sig == "AVOID":
            sig_color = "#B91C1C"
        elif "ACCUMULATE" in final_sig or "WATCH" in final_sig:
            sig_color = "#B45309"
        
        sig_style = ParagraphStyle(
            f"SigStyle_{i}",
            parent=tcc_style,
            textColor=colors.HexColor(sig_color),
            fontName="Helvetica-Bold"
        )

        rank_data.append([
            Paragraph(str(i), tcc_style),
            Paragraph(f"<b>{s.get('ticker', '')}</b>", tcc_style),
            Paragraph(s.get("sector") or "", tc_style),
            Paragraph(f'{s.get("fundamental_score")}/{s.get("fundamental_max")}', tcc_style),
            Paragraph(str(s.get("rsi", "–")), tcc_style),
            Paragraph(final_sig, sig_style)
        ])
        
    t = Table(rank_data, repeatRows=1, colWidths=[1.0*cm, 2.0*cm, 5.0*cm, 2.5*cm, 2.5*cm, 5.0*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # ---- Alerts ----
    if alerts:
        story.append(Paragraph("<b>Alerts Aktif</b>", styles["Heading2"]))
        alert_data = [[
            Paragraph("<b>Tipe</b>", th_style),
            Paragraph("<b>Ticker</b>", th_style),
            Paragraph("<b>Pesan</b>", th_style)
        ]]
        for a in alerts[:15]:
            alert_data.append([
                Paragraph(a.get("type", ""), tc_style),
                Paragraph(f"<b>{a.get('ticker', '')}</b>", tcc_style),
                Paragraph(a.get("message") or "", tc_style)
            ])
        t = Table(alert_data, repeatRows=1, colWidths=[3.5*cm, 2.5*cm, 12.0*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DC2626")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(t)

    # ---- Portfolio ----
    if portfolio:
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph("<b>Portfolio Saat Ini</b>", styles["Heading2"]))
        port_data = [[
            Paragraph("<b>Ticker</b>", th_style),
            Paragraph("<b>Beli</b>", th_style),
            Paragraph("<b>Sekarang</b>", th_style),
            Paragraph("<b>P/L</b>", th_style),
            Paragraph("<b>Aksi</b>", th_style)
        ]]
        for i, p in enumerate(portfolio, start=1):
            pnl_val = p.get("pnl", 0)
            pnl_text = f"{pnl_val:.2f}%"
            pnl_color = "#15803D" if pnl_val >= 0 else "#B91C1C"
            pnl_style = ParagraphStyle(
                f"PnlStyle_{i}",
                parent=tcc_style,
                textColor=colors.HexColor(pnl_color),
                fontName="Helvetica-Bold"
            )
            
            aksi_val = p.get("aksi", "")
            aksi_color = "black"
            if "SELL" in aksi_val:
                aksi_color = "#B91C1C"
            elif "BUY" in aksi_val:
                aksi_color = "#15803D"
            aksi_style = ParagraphStyle(
                f"AksiStyle_{i}",
                parent=tcc_style,
                textColor=colors.HexColor(aksi_color),
                fontName="Helvetica-Bold"
            )

            port_data.append([
                Paragraph(f"<b>{p.get('ticker', '')}</b>", tcc_style),
                Paragraph(f"{p.get('harga_beli', 0):,.0f}", tcc_style),
                Paragraph(f"{p.get('sekarang', 0):,.0f}", tcc_style),
                Paragraph(pnl_text, pnl_style),
                Paragraph(aksi_val, aksi_style)
            ])
        t = Table(port_data, repeatRows=1, colWidths=[2.5*cm, 3.5*cm, 3.5*cm, 4.5*cm, 4.0*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(t)

    doc.build(story)
    return out_path


# ----------------------------- BUILD SUMMARY ----------------------------- #

def build_summary(ranking: List[Dict], portfolio: List[Dict], alerts: List[Dict]) -> Dict:
    """Buat ringkasan dari data input."""
    avg_pnl = 0.0
    if portfolio:
        pnls = [p.get("pnl", 0) for p in portfolio]
        avg_pnl = sum(pnls) / len(pnls) if pnls else 0

    return {
        "total_stocks": len(ranking),
        "portfolio_count": len(portfolio),
        "avg_pnl": avg_pnl,
        "alert_count": len(alerts),
        "strong_buy_count": sum(1 for s in ranking
                                if s.get("final_signal") == "STRONG BUY"),
        "oversold_count": sum(1 for a in alerts
                              if a.get("type") == "OVERSOLD_GEM"),
    }
