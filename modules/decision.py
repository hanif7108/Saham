"""
Decision Center
===============
Mengagregasi sinyal dari portfolio + ranking + radar prospek + alerts
menjadi DAFTAR REKOMENDASI FINAL: BUY / SELL / HOLD dengan prioritas.

Filosofi:
  - User butuh keputusan yang JELAS tiap kali buka aplikasi.
  - Tiga bucket saja: BUY (peluang), SELL (urgent), HOLD (no action).
  - Setiap rekomendasi punya: confidence (HIGH/MEDIUM/LOW),
    urgency, reasoning ringkas, target harga (kalau ada).

Logika Decision Matrix
======================

UNTUK SAHAM DI PORTFOLIO (sudah dimiliki):
  P/L <= -7%                                 → SELL (CUT LOSS)        urgency=URGENT
  P/L >= +20%                                → SELL (TAKE PROFIT)     urgency=URGENT
  Fundamental AVOID                          → SELL (Fundamental rusak) urgency=HIGH
  Technical SELL + P/L < 0                   → SELL (Tren melemah)    urgency=MEDIUM
  Fundamental BUY + Technical BUY            → BUY MORE (Add)         confidence=HIGH
  Fundamental BUY + RSI <30                  → BUY MORE (Avg down)    confidence=HIGH
  Lainnya                                    → HOLD

UNTUK SAHAM TIDAK DI PORTFOLIO (watchlist):
  Fund>=4 + Tech BUY + Volume spike + RSI<60 → BUY (Strong Setup)     confidence=HIGH
  Fund>=4 + RSI<30 + Golden cross            → BUY (Oversold Gem)     confidence=HIGH
  Fund>=4 + Tech BUY                         → BUY (Quality + Momentum) confidence=MEDIUM
  Fund>=3 + RSI<30                           → BUY (Speculative bottom) confidence=LOW
  Lainnya                                    → HOLD (skip)

Output:
  {
    "buy":  [list of buy recommendations],
    "sell": [list of sell recommendations],
    "hold": int (count, tidak listed kecuali debug),
    "summary": {n_buy, n_sell, n_hold, top_pick, top_warning}
  }
"""

from typing import Dict, List, Optional

from .trading_mode import get_cut_loss_pct, get_take_profit_pct


# ----------------------------- Decision rules ----------------------------- #

def load_18langkah_current_step() -> Dict:
    """
    Membaca berkas Excel 18 Langkah Compounding Sakti secara dinamis
    untuk mendeteksi langkah aktif saat ini (berdasarkan baris pertama yang belum dicentang).
    """
    import os
    import openpyxl
    file_path = "/Users/hanif/Saham/PDS Trader Sakti/Trader Sakti Complete/KALKULATOR CUAN 2 MILIAR 18 Langkah Compounding Sakti.xlsx"
    step_info = {
        "step_no": 1,
        "growth_langkah": 0.2915,
        "saldo_awal": 10000000.0,
        "saldo_akhir": 12915496.65,
        "trades_per_langkah": 4,
        "growth_per_trade_butuh": 0.0661
    }
    try:
        if not os.path.exists(file_path):
            return step_info
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if "18 Langkah" in wb.sheetnames:
            sheet = wb["18 Langkah"]
            for r in range(4, 22):
                step_no = sheet.cell(r, 1).value
                growth_langkah = sheet.cell(r, 2).value
                saldo_awal = sheet.cell(r, 3).value
                saldo_akhir = sheet.cell(r, 4).value
                trades = sheet.cell(r, 5).value
                growth_trade = sheet.cell(r, 6).value
                checked_val = sheet.cell(r, 8).value # Col H
                
                is_checked = False
                if checked_val:
                    checked_str = str(checked_val).strip().lower()
                    if checked_str in ['☑', '✓', 'x', 'true', '1', 'v']:
                        is_checked = True
                
                if not is_checked and step_no is not None:
                    step_info["step_no"] = int(step_no)
                    step_info["growth_langkah"] = float(growth_langkah or 0.2915)
                    step_info["saldo_awal"] = float(saldo_awal or 10000000.0)
                    step_info["saldo_akhir"] = float(saldo_akhir or 12915496.65)
                    step_info["trades_per_langkah"] = int(trades or 4)
                    step_info["growth_per_trade_butuh"] = float(growth_trade or 0.0661)
                    break
    except Exception as e:
        print(f"[load_18langkah] Gagal membaca Excel: {e}")
    return step_info


def get_top5_ordered(ranking: List[Dict]) -> List[str]:
    """
    Hitung top 5 ticker syariah dengan kriteria yang sama dengan endpoint /api/top5_recommendations.
    """
    valid_candidates = []
    for s in (ranking or []):
        price = s.get("current_price") or 0.0
        fund_label = s.get("fundamental_label") or "AVOID"
        if price <= 0 or fund_label == "AVOID":
            continue
        
        fund_score = s.get("fundamental_score") or 0
        tech_score = s.get("technical_score") or 0
        combined_rank_score = (fund_score * 2) + tech_score
        
        valid_candidates.append({
            "ticker": s["ticker"],
            "combined_rank_score": combined_rank_score,
            "fundamental_score": fund_score,
            "technical_score": tech_score
        })
        
    valid_candidates.sort(
        key=lambda x: (
            x["combined_rank_score"],
            x["fundamental_score"],
            x["technical_score"]
        ),
        reverse=True
    )
    return [x["ticker"] for x in valid_candidates[:5]]


def decide_for_portfolio(
    p: Dict, 
    ranking_lookup: Dict[str, Dict], 
    top5_lookup: Dict[str, int],
    broker_cash: Optional[Dict[str, float]] = None,
    broker_holdings_map: Optional[Dict[str, str]] = None,
    step_info: Optional[Dict] = None
) -> Optional[Dict]:
    """
    Decide untuk saham di portfolio.
    """
    ticker = p.get("ticker")
    pnl = p.get("pnl", 0) or 0
    fund_score = p.get("fundamental_score", 0) or 0
    fund_max = p.get("fundamental_max", 0) or 0
    fund_label = p.get("aksi", "")  # current aksi field
    tech_signal = p.get("technical_signal", "HOLD")
    rsi = p.get("rsi") or 50
    ma_cross = p.get("ma_cross", "DEATH")

    # Tentukan broker tempat saham ini disimpan
    recommended_broker = "Bibit"
    if broker_holdings_map and ticker in broker_holdings_map:
        recommended_broker = broker_holdings_map[ticker]
    
    # Ambil kas RDN untuk broker bersangkutan
    cash_val = 0.0
    if broker_cash and recommended_broker in broker_cash:
        cash_val = float(broker_cash[recommended_broker])

    # Jika step_info terintegrasi, gunakan saldo awal langkah aktif sebagai basis modal
    base_capital_str = f"Kas RDN {recommended_broker}"
    base_capital_val = cash_val
    if step_info and "saldo_awal" in step_info:
        base_capital_val = step_info["saldo_awal"]
        base_capital_str = f"Modal Langkah {step_info['step_no']}"

    # --- SELL conditions (priority order) ---

    cl_threshold = get_cut_loss_pct()
    tp_threshold = get_take_profit_pct()
    if pnl <= cl_threshold:
        return {
            "action": "SELL",
            "ticker": ticker,
            "urgency": "URGENT",
            "confidence": "HIGH",
            "context": "PORTFOLIO",
            "reason": f"CUT LOSS — P/L {pnl}% sudah ≤ {cl_threshold}%",
            "details": {"pnl": pnl, "rule": "cut_loss_7pct"},
            "next_step": f"Jual di {recommended_broker} hari ini untuk batasi kerugian."
        }

    if pnl >= tp_threshold:
        return {
            "action": "SELL",
            "ticker": ticker,
            "urgency": "URGENT",
            "confidence": "HIGH",
            "context": "PORTFOLIO",
            "reason": f"TAKE PROFIT — P/L +{pnl}% sudah ≥ {tp_threshold}%",
            "details": {"pnl": pnl, "rule": "take_profit_20pct"},
            "next_step": f"Realisasikan profit di {recommended_broker} (boleh sebagian dulu)."
        }

    # Lihat ranking lookup untuk fundamental label terbaru
    rank_info = ranking_lookup.get(ticker, {})
    fund_label_real = rank_info.get("fundamental_label", "HOLD")

    # Jalur 7% pattern detection
    jalur7 = rank_info.get("jalur7", {}) or {}
    jalur7_signals = jalur7.get("signals", {}) if isinstance(jalur7, dict) else {}
    
    breaker_active = jalur7_signals.get("breaker", {}).get("pass", False) if isinstance(jalur7_signals, dict) else False
    kandang_kuda_active = jalur7_signals.get("kandang_kuda", {}).get("pass", False) if isinstance(jalur7_signals, dict) else False
    shadow_active = jalur7_signals.get("shadow", {}).get("pass", False) if isinstance(jalur7_signals, dict) else False
    swing_line_active = jalur7_signals.get("swing_line", {}).get("pass", False) if isinstance(jalur7_signals, dict) else False

    active_patterns = []
    if breaker_active:
        active_patterns.append("Breaker")
    if kandang_kuda_active:
        active_patterns.append("Kandang Kuda")
    if shadow_active:
        active_patterns.append("Shadow Rejection")
    if swing_line_active:
        active_patterns.append("Swing Line")

    jalur7_suffix = f" [Jalur 7%: {', '.join(active_patterns)} Aktif]" if active_patterns else ""

    if fund_label_real == "AVOID":
        return {
            "action": "SELL",
            "ticker": ticker,
            "urgency": "HIGH",
            "confidence": "MEDIUM",
            "context": "PORTFOLIO",
            "reason": f"Fundamental memburuk → AVOID (skor {fund_score}/{fund_max})",
            "details": {"pnl": pnl, "score": fund_score},
            "next_step": f"Eksit perlahan di {recommended_broker} walau belum kena cut-loss."
        }

    if tech_signal == "SELL" and pnl < 0 and ma_cross == "DEATH":
        return {
            "action": "SELL",
            "ticker": ticker,
            "urgency": "MEDIUM",
            "confidence": "MEDIUM",
            "context": "PORTFOLIO",
            "reason": f"Tech SELL + Death cross + P/L {pnl}% (sudah merah)",
            "details": {"pnl": pnl, "rsi": rsi, "ma_cross": ma_cross},
            "next_step": f"Pertimbangkan exit di {recommended_broker} jika besok masih melemah."
        }

    # --- BUY MORE conditions ---

    # Alokasi persentase modal/kas RDN
    def _alloc_desc(pct):
        if base_capital_val > 0:
            allocated = base_capital_val * pct / 100
            return f"Alokasikan {pct}% dari {base_capital_str} (±Rp {allocated:,.0f})"
        else:
            return f"Alokasikan {pct}% dari {base_capital_str} (Isi saldo terlebih dahulu)"

    # Base target profit (untuk expected_return_pct) — sama dengan watchlist
    base_target_pct_p = 7.0
    if step_info and "growth_per_trade_butuh" in step_info:
        base_target_pct_p = float(step_info["growth_per_trade_butuh"]) * 100

    if fund_score >= 4 and tech_signal == "BUY" and pnl >= -3:
        return {
            "action": "BUY",
            "ticker": ticker,
            "urgency": "MEDIUM",
            "confidence": "HIGH",
            "context": "PORTFOLIO_ADD",
            "reason": f"Add posisi — Fund {fund_score}/{fund_max} + Tech BUY{jalur7_suffix}",
            "details": {
                "pnl": pnl, "score": fund_score, "rsi": rsi, "jalur7": jalur7,
                "expected_return_pct": round(base_target_pct_p * 1.2, 2),
            },
            "next_step": f"Beli di {recommended_broker}. {_alloc_desc(10)} untuk tambah posisi."
        }

    if fund_score >= 4 and rsi < 30:
        return {
            "action": "BUY",
            "ticker": ticker,
            "urgency": "MEDIUM",
            "confidence": "HIGH",
            "context": "PORTFOLIO_AVERAGE_DOWN",
            "reason": f"Average down — Fund {fund_score}/{fund_max} + RSI {rsi:.1f} (oversold){jalur7_suffix}",
            "details": {
                "pnl": pnl, "score": fund_score, "rsi": rsi, "jalur7": jalur7,
                "expected_return_pct": round(base_target_pct_p * 1.4, 2),
            },
            "next_step": f"Beli di {recommended_broker}. {_alloc_desc(15)} untuk average down."
        }

    # Otherwise: HOLD (Hanya untuk saham yang sudah dimiliki di portfolio)
    rank_num = top5_lookup.get(ticker)
    reason_str = f"P/L {pnl:+.1f}% — Fundamental & teknikal stabil, pertahankan posisi."
    next_step_str = f"Terus simpan (HOLD) saham {ticker}."
    
    if rank_num is not None:
        reason_str = f"P/L {pnl:+.1f}% — Saham Top 5 (Pilihan Ke-{rank_num}), pertahankan posisi."
        next_step_str = f"Simpan (HOLD) pilihan utama ini."

    return {
        "action": "HOLD",
        "ticker": ticker,
        "urgency": "LOW",
        "confidence": "HIGH" if rank_num is not None else "MEDIUM",
        "context": "PORTFOLIO_HOLD",
        "reason": reason_str,
        "details": {"pnl": pnl, "score": fund_score, "rsi": rsi},
        "next_step": next_step_str
    }


def decide_for_watchlist(
    s: Dict, 
    prospects_set: set, 
    top5_lookup: Dict[str, int],
    broker_cash: Optional[Dict[str, float]] = None,
    step_info: Optional[Dict] = None
) -> Optional[Dict]:
    """
    Decide untuk saham yang TIDAK dimiliki (calon BUY).
    """
    ticker = s.get("ticker")
    fund_score = s.get("fundamental_score", 0) or 0
    fund_max = s.get("fundamental_max", 0) or 0
    fund_label = s.get("fundamental_label", "")
    tech_signal = s.get("technical_signal", "HOLD")
    rsi = s.get("rsi") or 50
    ma_cross = s.get("ma_cross", "DEATH")
    sector = s.get("sector", "")

    is_in_radar = ticker in prospects_set  # ada volume spike
    rank_num = top5_lookup.get(ticker)  # None jika bukan Top 5

    # Tentukan rekomendasi broker (Bibit atau Profit Anywhere) berdasarkan saldo kas RDN tertinggi
    recommended_broker = "Bibit"
    cash_val = 0.0
    if broker_cash:
        bibit_cash = float(broker_cash.get("Bibit", 0.0))
        profit_cash = float(broker_cash.get("Profit Anywhere", 0.0))
        if bibit_cash >= profit_cash:
            recommended_broker = "Bibit"
            cash_val = bibit_cash
        else:
            recommended_broker = "Profit Anywhere"
            cash_val = profit_cash

    # Jika step_info terintegrasi, gunakan saldo awal langkah aktif sebagai basis modal
    base_capital_str = f"Kas RDN {recommended_broker}"
    base_capital_val = cash_val
    if step_info and "saldo_awal" in step_info:
        base_capital_val = step_info["saldo_awal"]
        base_capital_str = f"Modal Langkah {step_info['step_no']}"

    # Helper untuk menghitung alokasi persentase
    def _alloc_desc(pct):
        if base_capital_val > 0:
            allocated = base_capital_val * pct / 100
            return f"Alokasikan {pct}% dari {base_capital_str} (±Rp {allocated:,.0f})"
        else:
            return f"Alokasikan {pct}% dari {base_capital_str} (Isi saldo terlebih dahulu)"

    # Target profit dinamis sesuai langkah compounding
    target_profit_str = "Target profit min 7%"
    base_target_pct = 7.0   # default kalau step_info kosong
    if step_info and "growth_per_trade_butuh" in step_info:
        base_target_pct = float(step_info["growth_per_trade_butuh"]) * 100
        target_profit_str = f"Target profit: {base_target_pct:.2f}% (Compounding Langkah {step_info['step_no']})"

    # Jalur 7% pattern detection
    jalur7 = s.get("jalur7", {}) or {}
    jalur7_signals = jalur7.get("signals", {}) if isinstance(jalur7, dict) else {}
    
    breaker_active = jalur7_signals.get("breaker", {}).get("pass", False) if isinstance(jalur7_signals, dict) else False
    kandang_kuda_active = jalur7_signals.get("kandang_kuda", {}).get("pass", False) if isinstance(jalur7_signals, dict) else False
    shadow_active = jalur7_signals.get("shadow", {}).get("pass", False) if isinstance(jalur7_signals, dict) else False
    swing_line_active = jalur7_signals.get("swing_line", {}).get("pass", False) if isinstance(jalur7_signals, dict) else False

    active_patterns = []
    if breaker_active:
        active_patterns.append("Breaker")
    if kandang_kuda_active:
        active_patterns.append("Kandang Kuda")
    if shadow_active:
        active_patterns.append("Shadow Rejection")
    if swing_line_active:
        active_patterns.append("Swing Line")

    jalur7_suffix = f" [Jalur 7%: {', '.join(active_patterns)} Aktif]" if active_patterns else ""

    # --- BUY conditions (priority order) ---

    # 1. KORELASI POSITIF: Jika masuk Top 5 Pilihan (Pilihan Utama)
    if rank_num is not None:
        if tech_signal == "BUY":
            return {
                "action": "BUY",
                "ticker": ticker,
                "urgency": "HIGH",
                "confidence": "HIGH",
                "context": "WATCHLIST",
                "reason": f"Top 5 Pilihan (Ke-{rank_num}) + Momentum — Fund {fund_score}/{fund_max} + Sinyal BUY{jalur7_suffix}",
                "details": {
                    "sector": sector, "rsi": rsi, "score": fund_score, "jalur7": jalur7,
                    "expected_return_pct": round(base_target_pct * 1.5, 2),
                },
                "next_step": f"Beli di {recommended_broker}. {_alloc_desc(20)} — {target_profit_str}."
            }
        elif tech_signal == "HOLD":
            return {
                "action": "BUY",
                "ticker": ticker,
                "urgency": "MEDIUM",
                "confidence": "HIGH",
                "context": "WATCHLIST",
                "reason": f"Top 5 Pilihan (Ke-{rank_num}) + DCA — Saham kualitas terbaik sedang konsolidasi/HOLD{jalur7_suffix}",
                "details": {
                    "sector": sector, "rsi": rsi, "score": fund_score, "jalur7": jalur7,
                    "expected_return_pct": round(base_target_pct * 1.1, 2),
                },
                "next_step": f"Beli di {recommended_broker}. {_alloc_desc(15)} secara bertahap."
            }

    # 2. Jalur 7% Setup (High Priority / High Confidence)
    if fund_score >= 4 and (breaker_active or kandang_kuda_active):
        patterns = " & ".join([p for p in ["Breaker", "Kandang Kuda"] if (p == "Breaker" and breaker_active) or (p == "Kandang Kuda" and kandang_kuda_active)])
        return {
            "action": "BUY",
            "ticker": ticker,
            "urgency": "HIGH",
            "confidence": "HIGH",
            "context": "WATCHLIST",
            "reason": f"Jalur 7% Setup ({patterns}) — Fund {fund_score}/{fund_max} + Pola aktif!{jalur7_suffix}",
            "details": {
                "sector": sector, "rsi": rsi, "score": fund_score, "jalur7": jalur7,
                "expected_return_pct": round(base_target_pct * 1.5, 2),
            },
            "next_step": f"Beli di {recommended_broker}. {_alloc_desc(15)} — {target_profit_str}."
        }

    # 3. Strong Setup
    if (fund_score >= 4 and tech_signal == "BUY"
            and is_in_radar and rsi < 60 and ma_cross == "GOLDEN"):
        return {
            "action": "BUY",
            "ticker": ticker,
            "urgency": "HIGH",
            "confidence": "HIGH",
            "context": "WATCHLIST",
            "reason": f"Strong Setup — Fund {fund_score}/{fund_max} + Tech BUY + Volume spike + Golden cross{jalur7_suffix}",
            "details": {
                "sector": sector, "rsi": rsi, "score": fund_score, "jalur7": jalur7,
                "expected_return_pct": round(base_target_pct * 1.3, 2),
            },
            "next_step": f"Beli di {recommended_broker}. {_alloc_desc(15)} — {target_profit_str}."
        }

    # 4. Oversold Gem
    if fund_score >= 4 and rsi < 30 and ma_cross == "GOLDEN":
        return {
            "action": "BUY",
            "ticker": ticker,
            "urgency": "MEDIUM",
            "confidence": "HIGH",
            "context": "WATCHLIST",
            "reason": f"Oversold Gem — Fund {fund_score}/{fund_max} ⭐⭐⭐⭐ + RSI {rsi:.1f} (bottom?) + masih Golden{jalur7_suffix}",
            "details": {
                "sector": sector, "rsi": rsi, "score": fund_score, "jalur7": jalur7,
                "expected_return_pct": round(base_target_pct * 1.4, 2),
            },
            "next_step": f"Beli di {recommended_broker}. {_alloc_desc(15)} secara bertahap (DCA) — {target_profit_str}."
        }

    # 5. Quality + Momentum
    if fund_score >= 4 and tech_signal == "BUY":
        return {
            "action": "BUY",
            "ticker": ticker,
            "urgency": "MEDIUM",
            "confidence": "MEDIUM",
            "context": "WATCHLIST",
            "reason": f"Quality + Momentum — Fund {fund_score}/{fund_max} + Tech BUY{jalur7_suffix}",
            "details": {
                "sector": sector, "rsi": rsi, "score": fund_score, "jalur7": jalur7,
                "expected_return_pct": round(base_target_pct * 1.0, 2),
            },
            "next_step": f"Beli di {recommended_broker}. {_alloc_desc(10)} — Tunggu koreksi sehat ke MA20. — {target_profit_str}."
        }

    # 6. Speculative bottom
    if fund_score >= 3 and rsi < 30:
        return {
            "action": "BUY",
            "ticker": ticker,
            "urgency": "LOW",
            "confidence": "LOW",
            "context": "WATCHLIST",
            "reason": f"Speculative bottom — Fund {fund_score}/{fund_max} + RSI {rsi:.1f}{jalur7_suffix}",
            "details": {
                "sector": sector, "rsi": rsi, "score": fund_score, "jalur7": jalur7,
                "expected_return_pct": round(base_target_pct * 0.7, 2),
            },
            "next_step": f"Beli di {recommended_broker}. {_alloc_desc(5)} (Posisi spekulatif kecil) — {target_profit_str}."
        }

    # Tidak ada sinyal → tidak listed (HOLD/skip)
    return None


# ----------------------------- Aggregator ----------------------------- #

def build_decisions(
    portfolio: List[Dict],
    ranking: List[Dict],
    prospects: List[Dict],
    market_status: Optional[Dict] = None,
    broker_cash: Optional[Dict[str, float]] = None,
    broker_holdings_map: Optional[Dict[str, str]] = None,
    step_info: Optional[Dict] = None
) -> Dict:
    """
    Bangun daftar rekomendasi final.

    Args:
        portfolio: hasil /api/portfolio
        ranking:   hasil /api/all_stocks (.stocks)
        prospects: hasil /api/prospects (untuk volume spike flag)
    """
    # Index ranking by ticker untuk fast lookup
    ranking_lookup = {r["ticker"]: r for r in (ranking or [])}
    portfolio_set = {p["ticker"] for p in (portfolio or [])}
    prospects_set = {p["ticker"] for p in (prospects or [])}

    # Hitung Top 5 secara dinamis
    top5_ordered = get_top5_ordered(ranking)
    top5_lookup = {ticker: idx + 1 for idx, ticker in enumerate(top5_ordered)}

    buy_list: List[Dict] = []
    sell_list: List[Dict] = []
    hold_list: List[Dict] = []

    # Iterate portfolio dulu (priority — saham yang sudah dimiliki)
    for p in (portfolio or []):
        decision = decide_for_portfolio(
            p, 
            ranking_lookup, 
            top5_lookup, 
            broker_cash=broker_cash, 
            broker_holdings_map=broker_holdings_map,
            step_info=step_info
        )
        if decision is None:
            continue
        if decision["action"] == "SELL":
            sell_list.append(decision)
        elif decision["action"] == "BUY":
            buy_list.append(decision)
        elif decision["action"] == "HOLD":
            hold_list.append(decision)

    # Iterate ranking untuk watchlist (yang BUKAN di portfolio)
    for s in (ranking or []):
        ticker = s.get("ticker")
        if ticker in portfolio_set:
            continue  # sudah dihandle di loop portfolio
        decision = decide_for_watchlist(
            s, 
            prospects_set, 
            top5_lookup, 
            broker_cash=broker_cash,
            step_info=step_info
        )
        if decision is None:
            continue
        if decision["action"] == "BUY":
            buy_list.append(decision)
        elif decision["action"] == "SELL":
            sell_list.append(decision)

    # Sort by urgency × confidence
    urgency_rank = {"URGENT": 4, "HIGH": 3, "MEDIUM": 2, "LOW": 1}
    confidence_rank = {"HIGH": 3, "MEDIUM": 2, "LOW": 1}

    def sort_key(d):
        return (urgency_rank.get(d["urgency"], 0),
                confidence_rank.get(d["confidence"], 0))

    sell_list.sort(key=sort_key, reverse=True)
    buy_list.sort(key=sort_key, reverse=True)
    hold_list.sort(key=sort_key, reverse=True)

    # Cek market direction (M dari CAN SLIM)
    is_uptrend = True
    if market_status and market_status.get("ihsg_close") and market_status.get("ihsg_ma50"):
        is_uptrend = market_status["ihsg_close"] > market_status["ihsg_ma50"]

    # Build summary
    summary = {
        "n_buy": len(buy_list),
        "n_sell": len(sell_list),
        "n_hold": len(hold_list),
        "top_pick": buy_list[0]["ticker"] if buy_list else None,
        "top_warning": sell_list[0]["ticker"] if sell_list else None,
        "market_uptrend": is_uptrend,
        "verdict": _global_verdict(buy_list, sell_list, is_uptrend, market_status)
    }

    return {
        "summary": summary,
        "sell": sell_list,
        "buy": buy_list,
        "hold": hold_list,
        "hold_count": len(hold_list)
    }


def _global_verdict(buy_list: List, sell_list: List,
                    is_uptrend: bool = True,
                    market_status: Optional[Dict] = None) -> str:
    """
    Verdict global untuk hari ini — narasi singkat 1 kalimat.
    Mengintegrasikan M (Market Direction) — kalau IHSG downtrend, suppress BUY.
    """
    n_buy = len(buy_list)
    n_sell = len(sell_list)

    # Market context string
    market_ctx = ""
    if market_status and market_status.get("ihsg_close") and market_status.get("ihsg_ma50"):
        delta = (market_status["ihsg_close"] - market_status["ihsg_ma50"]) \
                / market_status["ihsg_ma50"] * 100
        market_ctx = f" (IHSG {delta:+.1f}% vs MA50)"

    # PRIORITAS 1: Urgent sells (cut-loss/take-profit) — selalu wajib action
    urgent_sells = [s for s in sell_list if s["urgency"] == "URGENT"]
    if urgent_sells:
        tickers = ", ".join(s["ticker"] for s in urgent_sells[:3])
        return f"⚠️ AKSI WAJIB: Jual {tickers} hari ini (cut-loss/take-profit)."

    # PRIORITAS 2: Market downtrend — TAHAN entry baru meskipun ada peluang
    if not is_uptrend:
        if n_sell > 0:
            return (f"🔴 IHSG DOWNTREND{market_ctx} — {n_sell} aksi jual prioritas. "
                    f"Stop entry baru, tunggu IHSG balik atas MA50.")
        return (f"🔴 IHSG DOWNTREND{market_ctx} — TAHAN entry baru. "
                f"Fokus jaga cash, pantau saja {n_buy} setup yang muncul.")

    # PRIORITAS 3: Uptrend + high confidence buys
    high_buys = [b for b in buy_list if b["urgency"] == "HIGH" and b["confidence"] == "HIGH"]
    if high_buys:
        tickers = ", ".join(b["ticker"] for b in high_buys[:3])
        return f"🎯 PELUANG KUAT{market_ctx}: Pertimbangkan beli {tickers}."

    if n_sell > 0 and n_buy == 0:
        return f"🟡 Pasar lemah{market_ctx} — fokus exit posisi merah, hindari beli baru."
    if n_buy > 0 and n_sell == 0:
        return f"🟢 Uptrend{market_ctx} — ada {n_buy} peluang beli (low priority)."
    if n_buy == 0 and n_sell == 0:
        return f"🟡 Tidak ada sinyal kuat{market_ctx} — HOLD semua, sabar."

    return f"⚖️ Mixed{market_ctx} — {n_sell} jual, {n_buy} beli, evaluasi case-by-case."
